# -*- coding: utf-8 -*-
import pymysql
import urllib
import copy
import json
import time
import sys
import threading
import urllib.request
from urllib.request import urlopen
from openpyxl import load_workbook
from openpyxl import Workbook
import locale
locale.setlocale(locale.LC_ALL,"en_US.UTF-8")

#f = open('Log '+time.strftime("%Y-%m-%d", time.localtime())+'.log', 'w', encoding="utf-8")
#sys.stdout = f
#sys.stderr = f

people_head = 8
people_heads = ['full_name', 'family_name', 'last_name', 'entity_id', 'position', 'phone','email', 'comment']


# mysql connector
class sqlpool():
    def __init__(self, user, password, host, database):

        self.user = user
        self.password = password
        self.host = host
        self.database = database
        self.conn = pymysql.connect(user=user, 
                                    password=password,
                                    host=host,
                                    database=database,
                                    ssl={'ssl': {'ca': '/var/www/html/BaltimoreCyberTrustRoot.crt.pem'}})
        self.cursor = self.conn.cursor()
        self.saved_name = {}
        self.saved_entity = {}
        self.lock = threading.Lock()

    # Initialize people name list and entity name list
    def get_all_name(self):
        self.lock.acquire()
        self.cursor.execute('SELECT full_name,entity_id, people_id FROM people')
        for n in self.cursor.fetchall():
            self.saved_name["".join(filter(str.isalnum, n[0].lower()))+' '+str(n[1])] = n[2]

        self.cursor.execute('SELECT entity_name,entity_id FROM entity')
        for n in self.cursor.fetchall():
            self.saved_entity["".join(filter(str.isalnum, n[0].lower()))] = n[1]
        self.lock.release()

    # Search for person by name
    def search_name(self, name, entity_id):
        self.lock.acquire()
        plain_name = "".join(filter(str.isalnum, name.lower()))+' '+str(entity_id)
        result = self.saved_name.get(plain_name)
        self.lock.release()
        return result
    
    # Get stored data by name
    def get_stored_name(self, id):
        self.lock.acquire()
        result = None
        while result==None:
            self.cursor.execute('SELECT * FROM people WHERE people_id='+str(id))
            result = self.cursor.fetchone()
        self.lock.release()
        return result

    # Add new person
    def add_people(self, item):
        self.lock.acquire()
        try:
            self.cursor.execute('INSERT INTO People(full_name, family_name, last_name, entity_id, position, phone, email, comment, modify_time) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,NOW())', item)
            self.conn.commit()
            id=0
            while id<1:
                self.cursor.execute('SELECT people_id FROM people WHERE full_name="'+item[0]+'" AND entity_id='+str(item[3]))
                id = self.cursor.fetchone()[0]
            plain_name="".join(filter(str.isalnum, item[0].lower()))+' '+str(item[3])
            self.saved_name[plain_name] = id
            self.lock.release()
            print('New person insert succeed. name: '+item[0]+' id: '+str(id))
            return id
        except pymysql.InternalError as e:
            self.lock.release()
            print(e)
            print('Person insert failed.')
            return None
        self.lock.release()
    
    # Search entity name
    def search_entity(self, name):
        self.lock.acquire()
        plain_name = "".join(filter(str.isalnum, name.lower()))
        result = self.saved_entity.get(plain_name)
        self.lock.release()
        return result

    # add new entity name
    def add_entity(self, name):
        self.lock.acquire()
        try:
            self.cursor.execute('INSERT INTO Entity(entity_name, modify_time) VALUES("'+name+'", NOW())')
            self.conn.commit()
            id=0
            while id<1:
                self.cursor.execute('SELECT entity_id FROM entity WHERE entity_name="'+name+'"')
                id = self.cursor.fetchone()[0]
            plain_name="".join(filter(str.isalnum, name.lower()))
            self.saved_entity[plain_name] = id
            self.lock.release()
            print('Entity name insert succeed. name: '+name)
            return id
        except pymysql.InternalError as e:
            self.lock.release
            print(e)
            print('Entity name insert failed.')
            return None
        self.lock.release

    def commit(self, sql):
        self.lock.acquire()
        try:
            self.cursor.execute(sql)
            self.conn.commit()
            self.lock.release()
        except pymysql.InternalError as e:
            self.lock.release()
            print(e)
        self.lock.release()

    def close(self):
        self.cursor.close()
        self.conn.close()


# Read data from excel
def read_excel(file,default,heads,comments,first_line,db):
    wb = load_workbook(filename=file)
    sheet1 = wb.worksheets[0]
    row_num = sheet1.max_row
    print('Total roe number: '+str(row_num))
    col_num = sheet1.max_column
    print('Total col number: '+str(col_num))
    insert_num = 0
    update_num = 0
    fail = []

    #fill all none block
    for r in range(1, row_num+1):
        for c in range(1, col_num+1):
            if sheet1.cell(r, c).value is None:
                sheet1.cell(r, c, '')
            #else:
                #sheet1.cell(r, c).value.replace('\\','').replace('"',"'")

    #recognize heads
    i=1
    while(-1< i <= col_num):
        if heads[0] < 1 and (str(sheet1.cell(1,i).value).lower().find('first') !=-1 or str(sheet1.cell(1,i).value).lower().find('family') !=-1):
            print('First name is col: '+str(i))
            heads[0]=i
        elif heads[1] < 1 and str(sheet1.cell(1,i).value).lower().find('mid') != -1:
            print('Mid name is col: '+str(i))
            heads[1]=i
        elif heads[2] < 1 and str(sheet1.cell(1,i).value).lower().find('last') !=-1:
            print('Last name is col: '+str(i))
            heads[2]=i
        elif heads[3] < 1 and (str(sheet1.cell(1,i).value).lower().find('full') !=-1 or str(sheet1.cell(1,i).value).lower().find('name') !=-1):
            print('Full name is col: '+str(i))
            heads[3]=i
        elif heads[4] < 1 and str(sheet1.cell(1,i).value).lower().find('company') !=-1 :
            print('Company is col: '+str(i))
            heads[4]=i
        elif heads[5] < 1 and str(sheet1.cell(1,i).value).lower().find('position') !=-1 :
            print('Position is col: '+str(i))
            heads[5]=i
        elif heads[6] < 1 and (str(sheet1.cell(1,i).value).lower().find('phone') !=-1 or str(sheet1.cell(1,i).value).lower().find('mobile') !=-1 or str(sheet1.cell(1,i).value).find('T') !=-1):
            print('Phone is col: '+str(i))
            heads[6]=i
        elif heads[7] < 1 and (str(sheet1.cell(1,i).value).lower().find('mail') !=-1 or str(sheet1.cell(1,i).value).find('E') !=-1):
            print('Email is col: '+str(i))
            heads[7]=i
        elif  i > heads[8]:
            print('Comments include col: '+str(i))
            comments.append(i)
            heads[8]=i
        i+=1
        
    # return if cannot find commany name
    if heads[0]<1 and heads[2]<1 and heads[3]<1:
        print('No name coloum was found. Please enter a default head set.')
        return [0, 0, 0, 0, fail]
    
    # Get company detail
    fail.append(people_heads)
    r = first_line
    while(-1< r <= row_num):
        print('Row '+str(r))
        detail = copy.copy(default)
        if heads[3] > 0 and str(sheet1.cell(r,heads[3]).value).strip()!='':
            detail[0] = str(sheet1.cell(r,heads[3]).value).strip()
        if heads[2] > 0 and str(sheet1.cell(r,heads[2]).value).strip()!='' :
            detail[2] = str(sheet1.cell(r,heads[2]).value).strip()
        if heads[0] > 0 and str(sheet1.cell(r,heads[0]).value).strip()!='' :
            if heads[1] > 0:
                detail[1] = str(sheet1.cell(r,heads[0]).value).strip()+' '+str(sheet1.cell(r,heads[1]).value).strip()
            else:
                detail[1] = str(sheet1.cell(r,heads[0]).value).strip()
        if detail[0] != None or (detail[1]!=None and detail[2]!=None):
            if heads[4] > 0 and str(sheet1.cell(r,heads[4]).value).strip()!='':
                if str(sheet1.cell(r,heads[4]).value).lower().find('pty') != -1:
                    detail[3] = str(sheet1.cell(r,heads[4]).value)[0:str(sheet1.cell(r,heads[4]).value).lower().find('pty')].strip()+' PTY LTD'
                elif str(sheet1.cell(r,heads[0]).value).lower().find('p/l') != -1:
                    detail[3] = str(sheet1.cell(r,heads[4]).value)[0:str(sheet1.cell(r,heads[4]).value).lower().find('p/')].strip()
                else:
                    detail[3] = str(sheet1.cell(r,heads[4]).value).strip()+' PTY LTD'
            if heads[5] > 0 and str(sheet1.cell(r,heads[5]).value).strip()!='':
                detail[4] = str(sheet1.cell(r,heads[5]).value).strip()
            if heads[6] > 0 and str(sheet1.cell(r,heads[6]).value).strip()!='':
                detail[5] = str(sheet1.cell(r,heads[6]).value).strip().split(',')[0].split('/')[0]
            if heads[7] > 0 and str(sheet1.cell(r,heads[7]).value).strip()!='':
                detail[6] = str(sheet1.cell(r,heads[7]).value).strip().split(',')[0].split('/')[0]
            if heads[8] > 0:
                comment=''
                for com in comments:
                    if str(sheet1.cell(r,com).value).strip() != '': comment+=' '+str(sheet1.cell(1,com).value).strip()+': '+str(sheet1.cell(r,com).value).strip()+' ---'
                detail[7] = comment.strip()
            if detail[0] != None and (detail[1]==None or detail[2]==None):
                detail[2] = detail[0].split(' ')[-1]
                detail[1] = detail[0].replace(detail[2], '').strip()
            if detail[0] == None and detail[1]!=None and detail[2]!=None:
                detail[0] = detail[1] + ' ' + detail[2]
            if detail[0] != None and detail[1]!=None and detail[2]!=None:
                success = database_update(detail,db)
                if success == 0: insert_num+=1
                elif success == 1: update_num+=1
                elif success == 2: fail.append(detail)
            else:
                fail.append(detail)
                print('This person does not have full name. detail:')
                print(detail)
            print('-----')
        r+=1
    wb.close()
    return [row_num, insert_num, update_num, len(fail), fail]


# Check and update company detail
def database_update(item,db):
    
    if item[3]!=None:
        entity_id = db.search_entity(item[3])
        if entity_id == None:
            entity_id = db.add_entity(item[3])
        item[3] = db.search_entity(item[3])
    exist_name_id = db.search_name(item[0],item[3])
    if exist_name_id != None:
        print('Stored data matches.')

        olddata=db.get_stored_name(exist_name_id)
        sql='UPDATE People SET '
        i=4
        while(i < people_head):
            if item[i]!=None and olddata[i+1]==None:
                sql += people_heads[i]+'="'+item[i]+'", '
            i+=1
        if sql != 'UPDATE People SET ':
            sql += 'modify_time=NOW() WHERE people_id='+str(exist_name_id)
            try:
                db.commit(sql)
                print('Update succeed.')
                return 1
            except:
                print('Update failed. sql commend: '+sql)
                return 2
        else:
            print('No new data need to be updated.')
            return -1
    else:
        print('No stored data matches.')
        if db.add_people(item)!= None:
            return 0
        else:
            print(item)
            return 2


def print_fial_excel(fail_list):
    wb=Workbook(write_only=True)
    ws=wb.create_sheet('Fail list')
    for ele in fail_list:
        ws.append(ele)
    save_path=r'D:\excel\people fail list '+time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())+'.xlsx'
    wb.save(save_path)


if __name__=='__main__':
    start = time.time()

    #set default data
    file = 
    default = [None # 'full_name'
               ,None # 'family_name'
               ,None # 'last_name'
               ,None # 'entity_id'
               ,None # 'position'
               ,None # 'phone'
               ,None # 'email'
               ,None] # 'comment'

    heads = [0 # first name
             ,0 # mid name
             ,0 # last name
             ,0 # full name
             ,0 # company
             ,0 # position
             ,0 # phone
             ,0 # email
             ,0] # comment please set to last col
    comments = [] # Set useful comment col
    start_line = 2

    user='root'# 
    password=
    host='localhost'# 
    database=

    try:
        db = sqlpool(user,password,host,database)
        db.conn
        db.get_all_name()
        print('Connect database succeed.')
    except pymysql.Error as err:
        print(err)
        print('Connect database failed.')

    people = read_excel(file,default,heads,comments,start_line,db)
    print(str(people[0]-start_line)+' data have been read. '+str(people[1])+' of them have been insert.'+str(people[2])+' of them have been updated.'+str(people[3]-1)+' of them are unknow data.')
    print_fial_excel(people[4])

    db.close()
    
    time_elapsed = time.time() - start
    print('Code takes {:.0f}m {:.0f}s'.format(time_elapsed // 60, time_elapsed % 60))
