# -*- coding: utf-8 -*-
import pymysql
import urllib
import copy
import json
import time
import sys
import threading
import urllib.request
from urllib.request import urlopen
from openpyxl import load_workbook
from openpyxl import Workbook
import locale
locale.setlocale(locale.LC_ALL,"en_US.UTF-8")

#f = open('Log '+time.strftime("%Y-%m-%d", time.localtime())+'.log', 'w', encoding="utf-8")
#sys.stdout = f
#sys.stderr = f

company_head = 20
company_heads = ['company_name', 'category', 'entity_id', 'address_line1', 'address_line2', 'address_line3', 'sub', 'state', 'postcode', 'lat', 'lng', 'phone', 'fax', 'website','email', 'reg_number', 'contact_person','key_project','award', 'comment']


# mysql connector
class sqlpool():
    def __init__(self, user, password, host, database):

        self.user = user
        self.password = password
        self.host = host
        self.database = database
        self.conn = pymysql.connect(user=user, 
                                    password=password,
                                    host=host,
                                    database=database)
                                     #ssl={'ssl': {'ca': '/var/www/html/BaltimoreCyberTrustRoot.crt.pem'}}
        self.cursor = self.conn.cursor()
        self.saved_name = {}
        self.saved_entity = {}
        self.lock = threading.Lock()

    # Initialize company name list and entity name list
    def get_all_name(self):
        self.lock.acquire()
        self.cursor.execute('SELECT company_name,postcode,company_id FROM company')
        for n in self.cursor.fetchall():
            self.saved_name["".join(filter(str.isalnum, n[0].lower()))+' '+n[1]] = n[2]

        self.cursor.execute('SELECT entity_name,entity_id FROM entity')
        for n in self.cursor.fetchall():
            self.saved_entity["".join(filter(str.isalnum, n[0].lower()))] = n[1]
        self.lock.release()

    # Search for company by name
    def search_name(self, name, postcode):
        self.lock.acquire()
        plain_name = "".join(filter(str.isalnum, name.lower()))+' '+postcode
        result = self.saved_name.get(plain_name)
        self.lock.release()
        return result
    
    # Get stored data by name
    def get_stored_name(self, id):
        self.lock.acquire()
        result = None
        while result==None:
            self.cursor.execute('SELECT * FROM company WHERE company_id='+str(id))
            result = self.cursor.fetchone()
        self.lock.release()
        return result

    # Add new company
    def add_company(self, item):
        self.lock.acquire()
        try:
            self.cursor.execute('INSERT INTO Company(company_name, category, entity_id, address_line1, address_line2, address_line3, sub, state, postcode, lat, lng, phone, fax, website, email, reg_number, contact_person,key_project,award, comment, modify_time) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())', item)
            self.conn.commit()
            id=0
            while id<1:
                self.cursor.execute('SELECT company_id FROM company WHERE company_name="'+item[0]+'" AND postcode="'+item[8]+'"')
                id = self.cursor.fetchone()[0]
            plain_name="".join(filter(str.isalnum, item[0].lower()))+' '+item[8]
            self.saved_name[plain_name] = id
            self.lock.release()
            print('Company name insert succeed. name: '+item[0]+' id: '+str(id))
            return id
        except pymysql.InternalError as e:
            self.lock.release()
            print(e)
            print('Company name insert failed.')
            return None
        self.lock.release()
    
    # Search entity name
    def search_entity(self, name):
        self.lock.acquire()
        plain_name = "".join(filter(str.isalnum, name.lower()))
        result = self.saved_entity.get(plain_name)
        self.lock.release()
        return result

    # add new entity name
    def add_entity(self, name):
        self.lock.acquire()
        try:
            self.cursor.execute('INSERT INTO Entity(entity_name, modify_time) VALUES("'+name+'", NOW())')
            self.conn.commit()
            id=0
            while id<1:
                self.cursor.execute('SELECT entity_id FROM entity WHERE entity_name="'+name+'"')
                id = self.cursor.fetchone()[0]
            plain_name="".join(filter(str.isalnum, name.lower()))
            self.saved_entity[plain_name] = id
            self.lock.release()
            print('Entity name insert succeed. name: '+name)
            return id
        except pymysql.InternalError as e:
            self.lock.release
            print(e)
            print('Entity name insert failed.')
            return None
        self.lock.release

    def commit(self, sql):
        self.lock.acquire()
        try:
            self.cursor.execute(sql)
            self.conn.commit()
            self.lock.release()
        except pymysql.InternalError as e:
            self.lock.release()
            print(e)
        self.lock.release()

    def close(self):
        self.cursor.close()
        self.conn.close()


# Use Google map getting correct address
def getGeoForAddress(address, name):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36"}
    url = 'https://maps.googleapis.com/maps/api/geocode/json?address=' + address.replace(' ','%20').encode('ascii', 'ignore').decode() + '|country:AU&key=[your key]'
    req = urllib.request.Request(url = url, headers = headers)
    response = urlopen(req).read()
    responseJson = json.loads(response)
    sub = ''
    political = ''
    postcode = ''
    try:
        addressbag = responseJson.get('results')[0]['address_components']
        i=0
        while(1):
            try:
                if addressbag[i].get('types').count('administrative_area_level_1') > 0:
                    political = addressbag[i]['short_name']
                elif addressbag[i].get('types').count('postal_code') > 0:
                    postcode = addressbag[i]['short_name']
                elif addressbag[i].get('types').count('locality') > 0:
                    sub = addressbag[i]['short_name']
                i+=1
            except:
                break
        lat = responseJson.get('results')[0]['geometry']['location']['lat']
        lng = responseJson.get('results')[0]['geometry']['location']['lng']
        lines = responseJson.get('results')[0]['formatted_address'].split(',')
        line1 = lines[0]
        try:   
            line2 = lines[1].strip()
        except:
            line2 = ''
        try:   
            line3 = lines[2].strip()
        except:
            line3 = ''
    except:
        print('Plan A no result.')
    

    # Plan B
    if postcode == '':
        print('Plan A failed.')
        url2 = 'https://maps.googleapis.com/maps/api/place/textsearch/json?query=' + name.replace(' ','%20').encode('ascii', 'ignore').decode() + '&country:AU&key=[your key]'
        req2 = urllib.request.Request(url = url2, headers = headers)
        response = urlopen(req2).read()
        responseJson = json.loads(response)
        try:
            addressbag = responseJson.get('results')[0]
        except:
            print('Plan B no result.')
            return ['','','','','','','','']
        lat = responseJson.get('results')[0]['geometry']['location']['lat']
        lng = responseJson.get('results')[0]['geometry']['location']['lng']
        lines = responseJson.get('results')[0]['formatted_address'].split(',')
        line1 = lines[0]
        try:   
            line2 = lines[1].strip()
        except:
            line2 = ''
        try:   
            line3 = lines[2].strip()
        except:
            line3 = ''
        sublist = lines[-2].split(' ')
        postcode = sublist[-1]
        political = sublist[-2]
        sub = lines[-2].replace(postcode,'').replace(political,'').strip()

    if postcode == '':
        print('Plan B failed.')
        return ['','','','','','','','']
    return [line1,line2,line3, sub, political,postcode,lat,lng]



# Read data from excel
def read_excel(file,default,heads,address,comments,first_line,db):
    wb = load_workbook(filename=file)
    sheet1 = wb.worksheets[0]
    row_num = sheet1.max_row
    print('Total roe number: '+str(row_num))
    col_num = sheet1.max_column
    print('Total col number: '+str(col_num))
    insert_num = 0
    update_num = 0
    fail = []

    #fill all none block
    for r in range(1, row_num+1):
        for c in range(1, col_num+1):
            if sheet1.cell(r, c).value is None:
                sheet1.cell(r, c, '')

    #recognize heads
    i=1
    while(-1< i <= col_num):
        if heads[0] < 1 and str(sheet1.cell(1,i).value).lower().find('name') != -1:
            print('Name is col: '+str(i))
            heads[0]=i
        elif heads[1] < 1 and (str(sheet1.cell(1,i).value).lower().find('category') !=-1 or str(sheet1.cell(1,i).value).lower().find('type') !=-1):
            print('Category is col: '+str(i))
            heads[1]=i
        elif heads[2] < 1 and str(sheet1.cell(1,i).value).lower().find('entity') !=-1 :
            print('Entity is col: '+str(i))
            heads[2]=i
        elif heads[4] < 1 and (str(sheet1.cell(1,i).value).lower().find('phone') !=-1 or str(sheet1.cell(1,i).value).lower().find('mobile') !=-1 or str(sheet1.cell(1,i).value).find('T') !=-1):
            print('Phone is col: '+str(i))
            heads[4]=i
        elif heads[5] < 1 and (str(sheet1.cell(1,i).value).lower().find('fax') !=-1 or str(sheet1.cell(1,i).value).find('F') !=-1):
            print('Fax is col: '+str(i))
            heads[5]=i
        elif  heads[6] < 1 and (str(sheet1.cell(1,i).value).lower().find('web') !=-1 or str(sheet1.cell(1,i).value).lower().find('site') !=-1 or str(sheet1.cell(1,i).value).lower().find('url') !=-1 or str(sheet1.cell(1,i).value).find('W') !=-1):
            print('Web is col: '+str(i))
            heads[6]=i
        elif heads[7] < 1 and (str(sheet1.cell(1,i).value).lower().find('mail') !=-1 or str(sheet1.cell(1,i).value).find('E') !=-1):
            print('Email is col: '+str(i))
            heads[7]=i
        elif heads[8] < 1 and str(sheet1.cell(1,i).value).lower().find('reg') !=-1:
            print('Reg number is col: '+str(i))
            heads[8]=i
        elif heads[9] < 1 and str(sheet1.cell(1,i).value).lower().find('contact') !=-1:
            print('Contact is col: '+str(i))
            heads[9]=i
        elif heads[10] < 1 and str(sheet1.cell(1,i).value).lower().find('project') !=-1:
            print('Key project is col: '+str(i))
            heads[10]=i
        elif heads[11] < 1 and (str(sheet1.cell(1,i).value).lower().find('award') !=-1 or str(sheet1.cell(1,i).value).lower().find('prize') !=-1):
            print('Award is col: '+str(i))
            heads[11]=i
        elif heads[13] < 1 and str(sheet1.cell(1,i).value).lower().find('state') !=-1:
            print('State is col: '+str(i))
            address.append(i)
            heads[13]=i
            heads[3]=i
        elif heads[14] < 1 and str(sheet1.cell(1,i).value).lower().find('post') !=-1:
            print('Postcode is col: '+str(i))
            address.append(i)
            heads[14]=i
            heads[3]=i
        elif i > heads[3] and (str(sheet1.cell(1,i).value).lower().find('address') !=-1 or str(sheet1.cell(1,i).value).lower().find('street') !=-1  or str(sheet1.cell(1,i).value).lower().find('city') !=-1  or str(sheet1.cell(1,i).value).lower().find('sub') !=-1 or str(sheet1.cell(1,i).value).lower().find('area') !=-1):
            print('Address include col: '+str(i))
            address.append(i)
            heads[3]=i
        elif  i > heads[12]:
            print('Comments include col: '+str(i))
            comments.append(i)
            heads[12]=i
        i+=1
        
    # return if cannot find commany name
    if heads[0]<1:
        print('No company name coloum was found. Please enter a default head set.')
        return [0, 0, 0, 0, fail]
    
    # Get company detail
    fail.append(company_heads)
    r = first_line
    while(-1< r <= row_num):
        print('Row '+str(r)+' , Company name: '+str(sheet1.cell(r,heads[0]).value).strip())
        if str(sheet1.cell(r,heads[0]).value).strip() != '':
            detail = copy.copy(default)
            if str(sheet1.cell(r,heads[0]).value).lower().find('pty') != -1:
                detail[0] = str(sheet1.cell(r,heads[0]).value)[0:str(sheet1.cell(r,heads[0]).value).lower().find('pty')].replace('"','').strip()
                detail[2] = detail[0]+' PTY LTD'
            elif str(sheet1.cell(r,heads[0]).value).lower().find('p/') != -1:
                detail[0] = str(sheet1.cell(r,heads[0]).value)[0:str(sheet1.cell(r,heads[0]).value).lower().find('p/')].replace('"','').strip()
                detail[2] = detail[0]+' PTY LTD'
            else:
                detail[0] = str(sheet1.cell(r,heads[0]).value).replace('"','').strip()
            if heads[1] > 0 and str(sheet1.cell(r,heads[1]).value).strip()!='':
                detail[1] = str(sheet1.cell(r,heads[1]).value).replace('"','').strip()
            if head[2]>0  and str(sheet1.cell(r,heads[2]).value).strip()!='':
                if str(sheet1.cell(r,heads[2]).value).lower().find('pty') != -1:
                    detail[2] = str(sheet1.cell(r,heads[2]).value)[0:str(sheet1.cell(r,heads[2]).value).lower().find('pty')].replace('"','').strip()+' PTY LTD'
                elif str(sheet1.cell(r,heads[2]).value).lower().find('p/') != -1:
                    detail[2] = str(sheet1.cell(r,heads[2]).value)[0:str(sheet1.cell(r,heads[2]).value).lower().find('p/')].replace('"','').strip()+' PTY LTD'
                else:
                    detail[2] = str(sheet1.cell(r,heads[2]).value).replace('"','').strip()+' PTY LTD'
            if heads[3] > 0:
                addressline=''
                for c in address:
                    addressline+=str(sheet1.cell(r,c).value).strip()+' '
                if detail[2] == None:
                    addressline+=detail[0]
                    address2 = detail[0]+' '+"".join(filter(str.isalpha, str(sheet1.cell(r,heads[13]).value)))+' '+"".join(filter(str.isdigit, str(sheet1.cell(r,heads[14]).value)))
                else:
                    addressline+=detail[2]
                    address2 = detail[2]+' '+"".join(filter(str.isalpha, str(sheet1.cell(r,heads[13]).value)))+' '+"".join(filter(str.isdigit, str(sheet1.cell(r,heads[14]).value)))
                addressbag = getGeoForAddress(addressline.replace('/','%20').strip(), address2.replace('/','%20').strip())
                if addressbag[0] != '': detail[3] = addressbag[0]
                if addressbag[1] != '': detail[4] = addressbag[1]
                if addressbag[2] != '': detail[5] = addressbag[2]
                if addressbag[3] != '': detail[6] = addressbag[3]
                if addressbag[4] != '': 
                    detail[7] = addressbag[4]
                    detail[8] = addressbag[5]
                    detail[9] = str(addressbag[6])
                    detail[10] = str(addressbag[7])
            if heads[4] > 0 and str(sheet1.cell(r,heads[4]).value).strip()!='':
                detail[11] = str(sheet1.cell(r,heads[4]).value).strip().split(',')[0].split('/')[0].split(';')[0].replace('"','')
            if heads[5] > 0 and str(sheet1.cell(r,heads[5]).value).strip()!='':
                detail[12] = str(sheet1.cell(r,heads[5]).value).strip().split(',')[0].split('/')[0].split(';')[0].replace('"','')
            if heads[6] > 0 and str(sheet1.cell(r,heads[6]).value).strip()!='':
                detail[13] = str(sheet1.cell(r,heads[6]).value).strip().replace('"','')
            if heads[7] > 0 and str(sheet1.cell(r,heads[7]).value).strip()!='':
                detail[14] = str(sheet1.cell(r,heads[7]).value).strip().split(',')[0].split('/')[0].split(';')[0].replace('"','')
            if heads[8] > 0 and str(sheet1.cell(r,heads[8]).value).strip()!='':
                detail[15] = str(sheet1.cell(r,heads[8]).value).strip().replace('"','')
            if heads[9] > 0 and str(sheet1.cell(r,heads[9]).value).strip()!='':
                detail[16] = str(sheet1.cell(r,heads[9]).value).strip().replace('"','')
            if heads[10] > 0 and str(sheet1.cell(r,heads[10]).value).strip()!='':
                detail[17] = str(sheet1.cell(r,heads[10]).value).replace('"','').strip().replace('"','')
            if heads[11] > 0 and str(sheet1.cell(r,heads[11]).value).strip()!='':
                detail[18] = str(sheet1.cell(r,heads[11]).value).replace('"','').strip().replace('"','')
            if heads[12] > 0:
                comment=''
                for com in comments:
                    if str(sheet1.cell(r,com).value).strip() != '': comment+=' '+str(sheet1.cell(1,com).value).strip()+': '+str(sheet1.cell(r,com).value).strip()+' ---'
                detail[19] = comment.replace('"','').strip()
            if detail[8] != None: 
                success = database_update(detail,db)
                if success == 0: insert_num+=1
                elif success == 1: update_num+=1
                elif success == 2: fail.append(detail)
            else:
                if detail[7] == None and sheet1.cell(r,heads[13]).value != '': 
                    detail[7] = "".join(filter(str.isalpha, str(sheet1.cell(r,heads[13]).value)))
                if detail[8] == None and sheet1.cell(r,heads[14]).value != '': 
                    detail[8] = "".join(filter(str.isdigit, str(sheet1.cell(r,heads[14]).value)))
                    print('This company is not recognized by google, please check it later. detail:')
                else:
                    detail[8]='0000'
                    print('Company with postcode 0000 is not trustful, please check it later. detail:')
                fail.append(detail)
                print(detail)
                success = database_update(detail,db)
                if success == 0: insert_num+=1
                elif success == 1: update_num+=1
            print('-----')
        r+=1
    wb.close()
    return [row_num, insert_num, update_num, len(fail), fail]


# Check and update company detail
def database_update(item,db):
    
    print(item[0],item[8])
    exist_name_id = db.search_name(item[0],item[8])
    if exist_name_id != None:
        print('Stored data matches.')

        olddata=db.get_stored_name(exist_name_id)
        sql='UPDATE Company SET '
        if item[1] != None and olddata[2]=='un':
            sql += company_heads[1]+'="'+item[1]+'", '
        if item[2] != None and olddata[3]==None:
            entity_id = db.search_entity(item[2])
            if entity_id == None:
                db.add_entity(item[2])
                entity_id = db.search_entity(item[2])
            if entity_id != None:
                sql += company_heads[2]+'='+str(entity_id)+', '
        i=3
        while(i < company_head):
            if item[i]!=None and olddata[i+1]==None:
                sql += company_heads[i]+'="'+item[i]+'", '
            i+=1
        if sql != 'UPDATE Company SET ':
            sql += 'modify_time=NOW() WHERE company_id='+str(exist_name_id)
            try:
                db.commit(sql)
                print('Update succeed.')
                return 1
            except:
                print('Update failed. sql commend: '+sql)
                return 2
        else:
            print('No new data need to be updated.')
            return -1
    else:
        print('No stored data matches.')
        if item[2]!=None:
            entity_id = db.search_entity(item[2])
            if entity_id == None:
                entity_id = db.add_entity(item[2])
            item[2] = db.search_entity(item[2])
        if db.add_company(item)!= None:
            return 0
        else:
            print(item)
            return 2


def print_fial_excel(fail_list):
    wb=Workbook(write_only=True)
    ws=wb.create_sheet('Fail list')
    for ele in fail_list:
        ws.append(ele)
    save_path=r'D:\excel\fail list '+time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())+'.xlsx'
    wb.save(save_path)


if __name__=='__main__':
    start = time.time()

    #set default data
    file = 'D:\excel\Builders_plumbers_20191118 09-20-38.xlsx'
    #file = 'D:\excel\builder list 2019-11-18.xlsx'
    default = [None # 'company_name'
               ,None # 'type'
               ,None # 'brand'
               ,None # 'address_line1'
               ,None # 'address_line2'
               ,None # 'address_line3'
               ,None # 'sub'
               ,None # 'state'
               ,None # 'post_code'
               ,None # 'lat'
               ,None # 'lng'
               ,None # 'phone'
               ,None # 'fax'
               ,None # 'website'
               ,None # 'email'
               ,None # 'reg_number'
               ,None # 'contact_person'
               ,None # 'key_project'
               ,None # 'awards'
               ,None] # 'comment'

    heads = [0 # name
             ,0 # type
             ,0 # brand
             ,0 # address please set to last col
             ,0 # phone
             ,0 # fax
             ,0 # web
             ,0 # email
             ,0 # reg
             ,0 # contact
             ,0 # project
             ,0 # award
             ,0 # comment please set to last col
             ,0 # state
             ,0] # post
    address = [] # Set useful address col
    comments = [] # Set useful comment col
    start_line = 2

    user='root' 
    password='claude@1698774'
    host='localhost' 
    database='[db name]'
    try:
        db = sqlpool(user,password,host,database)
        db.conn
        db.get_all_name()
        print('Connect database succeed.')
    except pymysql.Error as err:
        print(err)
        print('Connect database failed.')

    company = read_excel(file,default,heads,address,comments,start_line,db)
    print(str(company[0]-start_line)+' data have been read. '+str(company[1])+' of them have been insert.'+str(company[2])+' of them have been updated.'+str(company[3]-1)+' of them are unknow data.')
    print_fial_excel(company[4])

    db.close()
    
    time_elapsed = time.time() - start
    print('Code takes {:.0f}m {:.0f}s'.format(time_elapsed // 60, time_elapsed % 60))